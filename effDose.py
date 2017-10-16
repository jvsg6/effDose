#!/usr/bin/python2.7
# -*- coding: utf-8 -*-
#Модуль для подготовки таблиц с дозами в населенных пунктах, приведенных в файле np.dat.
#Реализует медленную (но более точную, чем createVariationalSeriesNP) интерполяцию и сложение сгруппированных модулем createVariationalSeriesFI бинарных сеток, отражающих 
#преобразование выходных данных ПС НОСТРА в более удобный для программной обработки формат. Строятся таблицы, отражающие значения целевых функций,
#полученных с заданным уровнем доверия (по умолчанию - 95 процентный доверительный интервал)
#запуск: 1. запускается модуль createVariationalSeriesFI 3. запускается данный модуль.
#!!!!!важен порядок следования  функционалов!!!!!
#Разработан: 01.06.2017 
#Автор: Киселев А.А.
#Последняя модификация: 05.10.2017 
import sys
import os

from math import cos,sin,radians,sqrt,fabs,acos,pi,degrees,floor,ceil
import struct

import openpyxl

pathToFIdir = "/home/alexey/tasks/turkey_akkuyu_npp_2017/VVER_TOI_scenario_2/work//FI/"
pathToFrtIdir = "/home/alexey/tasks/turkey_akkuyu_npp_2017/VVER_TOI_scenario_2/work/FrtI/"
pathToNPfile = "/home/alexey/tasks/turkey_akkuyu_npp_2017/np.dat"

pointsForAnalysis = []
arrayOfResultsForPoints = []

def find_element_in_list(element,list_element):
        try:
		index_element=list_element.index(element)
		return index_element
	except ValueError:
		return -1 

def find_substr_in_list(element,list_element):
	index_element = []
	for i, item in enumerate(list_element):
		if item.find(element) >= 0:
			index_element.append(i)
	return index_element

def readLineF(line):
	lst = line.rstrip()
	vals = filter(None,lst.split(";"))
	return [float(x) for x in vals]

def readLineS(line):
	lst = line.rstrip()
	vals = filter(None,lst.split(";"))
	return [str(x) for x in vals]

def make_sure_path_exists(path):
	try: 
		os.makedirs(path)
	except OSError:
		if not os.path.isdir(path):
			raise

def reaNPFromFile(pathToFile):
	inp = open(pathToFile,"r")
	lines = inp.readlines()
	inp.close()
	lines = [line for line in lines if line.find("#") < 0]
	pointsForAnalysis = []
	for line in lines:
		pointsForAnalysis.append([str(x.strip("\"")) if i == 0 else float(x) for (i,x) in enumerate(filter(None,line.split(",")))])
	return pointsForAnalysis

class GeoGrid:
	def __init__(self,countx,county,lonmin,latmin,dlon,dlat,lonmax,latmax):
		self.lonmin = lonmin
		self.lonmax = lonmax
		self.latmin = latmin
		self.latmax = latmax
		self.countx = countx
		self.county = county
		self.fmin   = 0.0
		self.fmax   = 0.0		
		self.dlon = dlon
		self.dlat = dlat
		self.data   = [0.0 for o in range(countx*county)]
	
	def findMin(self):
		if(len(self.data) == 0):
			return -1
		ret = self.data[0]
		for v in self.data:
			if(ret > v):
				ret = v
		return ret
		
	def findMax(self):
		if(len(self.data) == 0):
			return -1
		ret = self.data[0]
		for v in self.data:
			if(ret < v):
				ret = v
		return ret			

	def printASCIIGRDFile(self,filename):
		f = open(filename, 'wt')
		f.write("DSAA" + '\r\n')
		f.write(str(self.countx)+"\t"+str(self.county) + '\r\n')
		f.write('%0.12f' % (self.lonmin)+"\t"+'%0.12f' % (self.lonmax) + '\r\n')
		f.write('%0.12f' % (self.latmin)+"\t"+'%0.12f' % (self.latmax) + '\r\n')
		f.write('%0.12e' % self.findMin()+"\t"+'%0.12e' % self.findMax() + '\r\n')
		for j in range(0,self.county):
			for i in range(0,self.countx):
				f.write('%0.12e' % self.data[i*self.county+j]+'\t')
			f.write('\r\n')
		
		f.close()
		return

	def getValue(self,lon,lat):
		if(lon<self.lonmin or lon>self.lonmax):
			return 0.
		if(lat<self.latmin or lat>self.latmax):
			return 0.
		ci = int(floor((lon-self.lonmin)/self.dlon))
		cj = int(floor((lat-self.latmin)/self.dlat))
		ci1 = 0
		ci2 = 0
                cj1 = 0
		cj2 = 0
		
		if ci == self.countx-1:
			ci2 = ci
			ci1 = ci-1
		else:
			ci1 = ci
			ci2 = ci+1
	    
		if cj == self.county-1:
			cj2 = cj
			cj1 = cj-1
		else:
			cj1 = cj
			cj2 = cj+1
		
		f11 = self.data[ci1*self.county+cj1] #+ -
		f21 = self.data[ci2*self.county+cj1] #+ -

		f12 = self.data[ci1*self.county+cj2] #- +
		f22 = self.data[ci2*self.county+cj2] #- +
		
		
		x1 = self.lonmin+self.dlon*ci1
		x2 = self.lonmin+self.dlon*ci2
		y1 = self.latmin+self.dlat*cj1
		y2 = self.latmin+self.dlat*cj2
		
		fy1=(f21-f11)/(x2-x1)*lon+(f11-(f21-f11)/(x2-x1)*x1)

		fy2=(f22-f12)/(x2-x1)*lon+(f12-(f22-f12)/(x2-x1)*x1)

		v = (fy2-fy1)/(y2-y1)*lat+(fy1-(fy2-fy1)/(y2-y1)*y1)
		return  v
	
	def angleOf(self,dX,dY):
		dFi = 0.0
		dR = sqrt(fabs(dX*dX+dY*dY));
		if (fabs(dR) > 0):
			dFi  = acos(dX/dR);
		if (dX <= 0 and dY < 0):
			dFi  = 2.0*pi - dFi;
		if (dX >  0 and dY < 0):
			dFi  = 2.0*pi - dFi;
		return dFi

def prepareArrayForVarSeries():
	global pointsForAnalysis, arrayOfResultsForPoints
	d = pathToFIdir
	listOfFiles = [os.path.join(d,o) for o in os.listdir(d) if (os.path.isfile(os.path.join(d,o)) and os.path.join(d,o).find(".bin")>=0)]
	for k in range(len(listOfFiles)): 
		fileName = pathToFIdir+str("/f"+str(k)+".bin")
		inp = open(fileName, 'rb')
		datastr = inp.read(64)
		offsetStart = 64
		headerStr = list(struct.unpack('<%dd' % (len(datastr)/8), datastr))
		grid = GeoGrid(int(headerStr[0]),int(headerStr[1]),headerStr[2],headerStr[3],headerStr[4],headerStr[5],headerStr[6],headerStr[7])
		del grid.data
		countx = int(headerStr[0])
		county = int(headerStr[1])
		offsetForNextGrid = countx*county*4
		gridsCount = (os.stat(fileName).st_size-offsetStart)/offsetForNextGrid-1
		for gridId in range(gridsCount):
			offset = offsetStart + gridId*offsetForNextGrid
			inp.seek(offset)
			grid.data = struct.unpack('<%df' % (countx*county), inp.read(offsetForNextGrid))
			for pointId,point in enumerate(pointsForAnalysis):
				if gridId == 0:
					arrayOfResultsForPoints[pointId].append([])
				arrayOfResultsForPoints[pointId][k].append(grid.getValue(point[1],point[2]))
			del grid.data
		inp.close()
		#if k == 3:
		#	break
		print "done", str(float(k)/len(listOfFiles)*100.0)+" %"
		
	#вариант Асфандиярова (не совсем корректный)
	#for pointId,point in enumerate(pointsForAnalysis):
	#	for k in range(len(arrayOfResultsForPoints[pointId])):
	#		arrayOfResultsForPoints[pointId][k].sort()
	return

def sumGridForDoseNew(*args):
	if len(args) == 0:
		return None
	result = [[0.0 for j in range(len(arrayOfResultsForPoints[i][0]))] for i in range(len(pointsForAnalysis))]
	for arg in args:
		idx = int(arg[1:])
		for pointId in range(len(pointsForAnalysis)):
			for j in range(len(arrayOfResultsForPoints[pointId][idx])):
				result[pointId][j] = result[pointId][j] + arrayOfResultsForPoints[pointId][idx][j]
	return result

def maxGridForDoseNew(*args):
	if len(args) == 0:
		return None
	result = [[0.0 for j in range(len(arrayOfResultsForPoints[i][0]))] for i in range(len(pointsForAnalysis))]
	for arg in args:
		idx = int(arg[1:])
		for pointId in range(len(pointsForAnalysis)):
			for j in range(len(arrayOfResultsForPoints[pointId][idx])):
				if arrayOfResultsForPoints[pointId][idx][j] > result[pointId][j]:
					result[pointId][j] = arrayOfResultsForPoints[pointId][idx][j]
	return result

def sumGridForDose(*args):
	if len(args) == 0:
		return None
	result = [[0.0 for j in range(len(arrayOfResultsForPoints[i][0]))] for i in range(len(pointsForAnalysis))]
	for arg in args:
		for pointId in range(len(pointsForAnalysis)):
			for j in range(len(arg[pointId])):
				result[pointId][j] = result[pointId][j] + arg[pointId][j]
	for arg in args:
		del arg[:]
	return result
	
def maxGridForDose(*args):
	if len(args) == 0:
		return None
	result = [[0.0 for j in range(len(arrayOfResultsForPoints[i][0]))] for i in range(len(pointsForAnalysis))]
	for arg in args:
		for pointId in range(len(pointsForAnalysis)):
			for j in range(len(arg[pointId])):
				if arg[pointId][j] > result[pointId][j]:
					result[pointId][j] = arg[pointId][j]
	for arg in args:
		del arg[:]
	return result
	
def getValueForConfidence(pointId,array,confVal = 95.0):
	idx = int(ceil(float(len(array[pointId]))*confVal/100.0))-1
	tmpArr = array[pointId][:]
	tmpArr.sort()
	val = tmpArr[idx]
	del tmpArr
	return val

# Таблица 1 – Прогнозируемые дозы за счет внешнего облучения от облака и от выпадений на поверхность земли (взвешенные по ОБЭ дозы за первые 10 ч), Гр . scenario
def createTable1(scenario):
	_RED_MARROW = sumGridForDoseNew("f10","f11")
	_LUNGS = sumGridForDoseNew("f12","f13")
	_OVARIES = sumGridForDoseNew("f14","f15")
	_TESTES = sumGridForDoseNew("f16","f17")
	_THYROID = sumGridForDoseNew("f18","f19")
	_FOETUS = sumGridForDoseNew("f3","f7")
	
	wb = openpyxl.Workbook()
	ws = wb.worksheets[0]
	ws.title = u'Таблица 1'
	ws.merge_cells('A1:F1')
	ws.merge_cells('A2:A3')
	ws.merge_cells('B2:F2')
	
	ws['A1'] = u'Таблица 1 – Прогнозируемые дозы за счет внешнего облучения от облака и от выпадений на поверхность земли (взвешенные по ОБЭ дозы за первые 10 ч), Гр . scenario'.replace("scenario",scenario)
	ws['A2'] = u'Населенные пункты (долгота; широта)'
	ws['B2'] = u'Дозы на органы'
	
	ws['B3'] = u'красный костный мозг'
	ws['C3'] = u'легкие'
	ws['D3'] = u'гонады'
	ws['E3'] = u'щитовидная железа'
	ws['F3'] = u'плод'
	
	curPos = 4
	for pointId, point in enumerate(pointsForAnalysis):
		ws['A'+str(curPos)] = point[0]+" ("+ format(point[1], '.6f').replace('.', ',')+";"+format(point[2], '.6f').replace('.', ',')+")"
		ws['B'+str(curPos)] = format(getValueForConfidence(pointId,_RED_MARROW), '.2E').replace('.', ',')
		ws['C'+str(curPos)] = format(getValueForConfidence(pointId,_LUNGS), '.2E').replace('.', ',')
		ws['D'+str(curPos)] = format(max(getValueForConfidence(pointId,_OVARIES),getValueForConfidence(pointId,_TESTES)), '.2E').replace('.', ',')
		ws['E'+str(curPos)] = format(getValueForConfidence(pointId,_THYROID), '.2E').replace('.', ',')
		ws['F'+str(curPos)] = format(getValueForConfidence(pointId,_FOETUS), '.2E').replace('.', ',')
		curPos = curPos + 1
	
	wb.save(pathToFrtIdir+"/Table1.xlsx")
	
	del _RED_MARROW[:]
	del _LUNGS[:]
	del _OVARIES[:]
	del _TESTES[:]
	del _THYROID[:]
	del _FOETUS[:]
	
	return 

# Таблица 2 – Прогнозируемые дозы за счет: внутреннего облучения от ингаляционного поступления радионуклидов для взрослых (взвешенные по ОБЭ дозы по внутреннему пути облучения за период 30 суток), Гр . scenario
def createTable2(scenario):
	_RED_MARROW = maxGridForDoseNew("f37")
	_THYROID = maxGridForDoseNew("f53")
	_LUNGS = maxGridForDoseNew("f45")
	_UPPER_LARGE_INTESTINE_WALL = maxGridForDoseNew("f61")
	_LOWER_LARGE_INTESTINE_WALL = maxGridForDoseNew("f69")
	_FOETUS = maxGridForDoseNew("f2")
	
	wb = openpyxl.Workbook()
	ws = wb.worksheets[0]
	ws.title = u'Таблица 2'
	ws.merge_cells('A1:F1')
	ws.merge_cells('A2:A3')
	ws.merge_cells('B2:F2')
	
	ws['A1'] = u'Таблица 2 – Прогнозируемые дозы за счет: внутреннего облучения от ингаляционного поступления радионуклидов для взрослых (взвешенные по ОБЭ дозы по внутреннему пути облучения за период 30 суток), Гр . scenario'.replace("scenario",scenario)
	ws['A2'] = u'Населенные пункты (долгота; широта)'
	ws['B2'] = u'Дозы на органы (взрослые)'
	
	ws['B3'] = u'красный костный мозг'
	ws['C3'] = u'щитовидная железа'
	ws['D3'] = u'легкие'
	ws['E3'] = u'толстый кишечник'
	ws['F3'] = u'плод'
	
	curPos = 4
	for pointId, point in enumerate(pointsForAnalysis):
		ws['A'+str(curPos)] = point[0]+" ("+ format(point[1], '.6f').replace('.', ',')+";"+format(point[2], '.6f').replace('.', ',')+")"
		ws['B'+str(curPos)] = format(getValueForConfidence(pointId,_RED_MARROW), '.2E').replace('.', ',')
		ws['C'+str(curPos)] = format(getValueForConfidence(pointId,_THYROID), '.2E').replace('.', ',')
		ws['D'+str(curPos)] = format(getValueForConfidence(pointId,_LUNGS), '.2E').replace('.', ',')
		ws['E'+str(curPos)] = format(0.57*getValueForConfidence(pointId,_UPPER_LARGE_INTESTINE_WALL) + 0.43*getValueForConfidence(pointId,_LOWER_LARGE_INTESTINE_WALL) , '.2E').replace('.', ',')
		ws['F'+str(curPos)] = format(getValueForConfidence(pointId,_FOETUS), '.2E').replace('.', ',')
		curPos = curPos + 1
	
	wb.save(pathToFrtIdir+"/Table2.xlsx")
	
	del _RED_MARROW[:]
	del _THYROID[:]
	del _LUNGS[:]
	del _UPPER_LARGE_INTESTINE_WALL[:]
	del _LOWER_LARGE_INTESTINE_WALL[:]
	del _FOETUS[:]
	
	return

# Таблица 3 – Прогнозируемые дозы за счет: внутреннего облучения от ингаляционного поступления радионуклидов для детей (взвешенные по ОБЭ дозы по внутреннему пути облучения за период 30 суток), Гр . scenario
def createTable3(scenario):
	_RED_MARROW = maxGridForDoseNew("f32","f33","f34","f35","f36")
	_THYROID = maxGridForDoseNew("f48","f49","f50","f51","f52")
	_LUNGS = maxGridForDoseNew("f40","f41","f42","f43","f44")
	_UPPER_LARGE_INTESTINE_WALL = maxGridForDoseNew("f56","f57","f58","f59","f60")
	_LOWER_LARGE_INTESTINE_WALL = maxGridForDoseNew("f64","f65","f66","f67","f68")
	
	wb = openpyxl.Workbook()
	ws = wb.worksheets[0]
	ws.title = u'Таблица 3'
	ws.merge_cells('A1:E1')
	ws.merge_cells('A2:A3')
	ws.merge_cells('B2:E2')
	
	ws['A1'] = u'Таблица 3 – Прогнозируемые дозы за счет: внутреннего облучения от ингаляционного поступления радионуклидов для детей (взвешенные по ОБЭ дозы по внутреннему пути облучения за период 30 суток), Гр . scenario'.replace("scenario",scenario)
	ws['A2'] = u'Населенные пункты (долгота; широта)'
	ws['B2'] = u'Дозы на органы (дети)'
	
	ws['B3'] = u'красный костный мозг'
	ws['C3'] = u'щитовидная железа'
	ws['D3'] = u'легкие'
	ws['E3'] = u'толстый кишечник'
	
	curPos = 4
	for pointId, point in enumerate(pointsForAnalysis):
		ws['A'+str(curPos)] = point[0]+" ("+ format(point[1], '.6f').replace('.', ',')+";"+format(point[2], '.6f').replace('.', ',')+")"
		ws['B'+str(curPos)] = format(getValueForConfidence(pointId,_RED_MARROW), '.2E').replace('.', ',')
		ws['C'+str(curPos)] = format(getValueForConfidence(pointId,_THYROID), '.2E').replace('.', ',')
		ws['D'+str(curPos)] = format(getValueForConfidence(pointId,_LUNGS), '.2E').replace('.', ',')
		ws['E'+str(curPos)] = format(0.57*getValueForConfidence(pointId,_UPPER_LARGE_INTESTINE_WALL) + 0.43*getValueForConfidence(pointId,_LOWER_LARGE_INTESTINE_WALL) , '.2E').replace('.', ',')
		curPos = curPos + 1
	
	wb.save(pathToFrtIdir+"/Table3.xlsx")
	
	del _RED_MARROW[:]
	del _THYROID[:]
	del _LUNGS[:]
	del _UPPER_LARGE_INTESTINE_WALL[:]
	del _LOWER_LARGE_INTESTINE_WALL[:]
	
	return

# Таблица 4 – Прогнозируемые дозы за счет: внешнего облучения от облака и от выпадений на поверхность земли (взвешенные по ОБЭ дозы за первые 30 суток), Гр . scenario
def createTable4(scenario):
	_RED_MARROW = sumGridForDoseNew("f30","f31")
	_THYROID = sumGridForDoseNew("f46","f47")
	_LUNGS = sumGridForDoseNew("f38","f39")
	_UPPER_LARGE_INTESTINE_WALL = sumGridForDoseNew("f54","f55")
	_LOWER_LARGE_INTESTINE_WALL = sumGridForDoseNew("f62","f63")
	_FOETUS = sumGridForDoseNew("f3","f9")
	
	wb = openpyxl.Workbook()
	ws = wb.worksheets[0]
	ws.title = u'Таблица 4'
	ws.merge_cells('A1:F1')
	ws.merge_cells('A2:A3')
	ws.merge_cells('B2:F2')
	
	ws['A1'] = u'Таблица 4 – Прогнозируемые дозы за счет: внешнего облучения от облака и от выпадений на поверхность земли (взвешенные по ОБЭ дозы за первые 30 суток), Гр . scenario'.replace("scenario",scenario)
	ws['A2'] = u'Населенные пункты (долгота; широта)'
	ws['B2'] = u'Дозы на органы'
	
	ws['B3'] = u'красный костный мозг'
	ws['C3'] = u'щитовидная железа'
	ws['D3'] = u'легкие'
	ws['E3'] = u'толстый кишечник'
	ws['F3'] = u'плод'
	
	curPos = 4
	for pointId, point in enumerate(pointsForAnalysis):
		ws['A'+str(curPos)] = point[0]+" ("+ format(point[1], '.6f').replace('.', ',')+";"+format(point[2], '.6f').replace('.', ',')+")"
		ws['B'+str(curPos)] = format(getValueForConfidence(pointId,_RED_MARROW), '.2E').replace('.', ',')
		ws['C'+str(curPos)] = format(getValueForConfidence(pointId,_THYROID), '.2E').replace('.', ',')
		ws['D'+str(curPos)] = format(getValueForConfidence(pointId,_LUNGS), '.2E').replace('.', ',')
		ws['E'+str(curPos)] = format(0.57*getValueForConfidence(pointId,_UPPER_LARGE_INTESTINE_WALL) + 0.43*getValueForConfidence(pointId,_LOWER_LARGE_INTESTINE_WALL) , '.2E').replace('.', ',')
		ws['F'+str(curPos)] = format(getValueForConfidence(pointId,_FOETUS), '.2E').replace('.', ',')
		curPos = curPos + 1
	
	wb.save(pathToFrtIdir+"/Table4.xlsx")
	
	del _RED_MARROW[:]
	del _THYROID[:]
	del _LUNGS[:]
	del _UPPER_LARGE_INTESTINE_WALL[:]
	del _LOWER_LARGE_INTESTINE_WALL[:]
	del _FOETUS[:]
	
	return

# Таблица 5 – Прогнозируемые эквивалентные и эффективная дозы за счет: облучения от облака и от выпадений на поверхность земли и внутреннего облучения от ингаляционного поступления радионуклидов для взрослых (доза за 10 суток), Зв . scenario
def createTable5(scenario):
	_RED_MARROW = sumGridForDoseNew("f142","f143","f149")
	_THYROID = sumGridForDoseNew("f166","f167","f173")
	_LUNGS = sumGridForDoseNew("f150","f151","f157")
	_SMALL_INTESTINE_WALL = sumGridForDoseNew("f190","f191","f197")
	_SKIN = sumGridForDoseNew("f158","f159","f165")
	_TESTES = sumGridForDoseNew("f182","f183","f189")
	_OVARIES = sumGridForDoseNew("f174","f175","f181")
	_EFFECTIVE_DOSE = sumGridForDoseNew("f134","f135","f141")
	
	wb = openpyxl.Workbook()
	ws = wb.worksheets[0]
	ws.title = u'Таблица 5'
	ws.merge_cells('A1:H1')
	ws.merge_cells('A2:A3')
	ws.merge_cells('B2:H2')
	
	ws['A1'] = u'Таблица 5 – Прогнозируемые эквивалентные и эффективная дозы за счет: облучения от облака и от выпадений на поверхность земли и внутреннего облучения от ингаляционного поступления радионуклидов для взрослых (доза за 10 суток), Зв . scenario'.replace("scenario",scenario)
	ws['A2'] = u'Населенные пункты (долгота; широта)'
	ws['B2'] = u'Дозы на органы (взрослые)'
	
	ws['B3'] = u'красный костный мозг'
	ws['C3'] = u'щитовидная железа'
	ws['D3'] = u'легкие'
	ws['E3'] = u'тонкий кишечник'
	ws['F3'] = u'кожа'
	ws['G3'] = u'гонады'
	ws['H3'] = u'эффективная доза'
	
	curPos = 4
	for pointId, point in enumerate(pointsForAnalysis):
		ws['A'+str(curPos)] = point[0]+" ("+ format(point[1], '.6f').replace('.', ',')+";"+format(point[2], '.6f').replace('.', ',')+")"
		ws['B'+str(curPos)] = format(getValueForConfidence(pointId,_RED_MARROW), '.2E').replace('.', ',')
		ws['C'+str(curPos)] = format(getValueForConfidence(pointId,_THYROID), '.2E').replace('.', ',')
		ws['D'+str(curPos)] = format(getValueForConfidence(pointId,_LUNGS), '.2E').replace('.', ',')
		ws['E'+str(curPos)] = format(getValueForConfidence(pointId,_SMALL_INTESTINE_WALL) , '.2E').replace('.', ',')
		ws['F'+str(curPos)] = format(getValueForConfidence(pointId,_SKIN), '.2E').replace('.', ',')
		ws['G'+str(curPos)] = format(max(getValueForConfidence(pointId,_OVARIES),getValueForConfidence(pointId,_TESTES)), '.2E').replace('.', ',')
		ws['H'+str(curPos)] = format(getValueForConfidence(pointId,_EFFECTIVE_DOSE) , '.2E').replace('.', ',')
		curPos = curPos + 1
	
	wb.save(pathToFrtIdir+"/Table5.xlsx")
	
	del _RED_MARROW[:]
	del _THYROID[:]
	del _LUNGS[:]
	del _SMALL_INTESTINE_WALL[:]
	del _SKIN[:]
	del _TESTES[:]
	del _OVARIES[:]
	del _EFFECTIVE_DOSE[:]
	
	return

# Таблица 6 – Прогнозируемые эквивалентные и эффективная дозы за счет: облучения от облака и от выпадений на поверхность земли и внутреннего облучения от ингаляционного поступления радионуклидов для детей (доза за 10 суток), Зв . scenario
def createTable6(scenario):
	_RED_MARROW = sumGridForDose(sumGridForDoseNew("f142","f143"),maxGridForDoseNew("f144","f145","f146","f147","f148"))
	_THYROID = sumGridForDose(sumGridForDoseNew("f166","f167"),maxGridForDoseNew("f168","f169","f170","f171","f172"))
	_LUNGS = sumGridForDose(sumGridForDoseNew("f150","f151"),maxGridForDoseNew("f152","f153","f154","f155","f156"))
	_SMALL_INTESTINE_WALL = sumGridForDose(sumGridForDoseNew("f190","f191"),maxGridForDoseNew("f192","f193","f194","f195","f196"))
	_SKIN = sumGridForDose(sumGridForDoseNew("f158","f159"),maxGridForDoseNew("f160","f161","f162","f163","f164"))
	_TESTES = sumGridForDose(sumGridForDoseNew("f182","f183"),maxGridForDoseNew("f184","f185","f186","f187","f188"))
	_OVARIES = sumGridForDose(sumGridForDoseNew("f174","f175"),maxGridForDoseNew("f176","f177","f178","f179","f180"))
	_EFFECTIVE_DOSE = sumGridForDose(sumGridForDoseNew("f134","f135"),maxGridForDoseNew("f136","f137","f138","f139","f140"))
	
	wb = openpyxl.Workbook()
	ws = wb.worksheets[0]
	ws.title = u'Таблица 6'
	ws.merge_cells('A1:H1')
	ws.merge_cells('A2:A3')
	ws.merge_cells('B2:H2')
	
	ws['A1'] = u'Таблица 6 – Прогнозируемые эквивалентные и эффективная дозы за счет: облучения от облака и от выпадений на поверхность земли и внутреннего облучения от ингаляционного поступления радионуклидов для детей (доза за 10 суток), Зв . scenario'.replace("scenario",scenario)
	ws['A2'] = u'Населенные пункты (долгота; широта)'
	ws['B2'] = u'Дозы на органы (дети)'
	
	ws['B3'] = u'красный костный мозг'
	ws['C3'] = u'щитовидная железа'
	ws['D3'] = u'легкие'
	ws['E3'] = u'тонкий кишечник'
	ws['F3'] = u'кожа'
	ws['G3'] = u'гонады'
	ws['H3'] = u'эффективная доза'
	
	curPos = 4
	for pointId, point in enumerate(pointsForAnalysis):	
		ws['A'+str(curPos)] = point[0]+" ("+ format(point[1], '.6f').replace('.', ',')+";"+format(point[2], '.6f').replace('.', ',')+")"
		ws['B'+str(curPos)] = format(getValueForConfidence(pointId,_RED_MARROW), '.2E').replace('.', ',')
		ws['C'+str(curPos)] = format(getValueForConfidence(pointId,_THYROID), '.2E').replace('.', ',')
		ws['D'+str(curPos)] = format(getValueForConfidence(pointId,_LUNGS), '.2E').replace('.', ',')
		ws['E'+str(curPos)] = format(getValueForConfidence(pointId,_SMALL_INTESTINE_WALL) , '.2E').replace('.', ',')
		ws['F'+str(curPos)] = format(getValueForConfidence(pointId,_SKIN), '.2E').replace('.', ',')
		ws['G'+str(curPos)] = format(max(getValueForConfidence(pointId,_OVARIES),getValueForConfidence(pointId,_TESTES)), '.2E').replace('.', ',')
		ws['H'+str(curPos)] = format(getValueForConfidence(pointId,_EFFECTIVE_DOSE) , '.2E').replace('.', ',')
		curPos = curPos + 1
	
	wb.save(pathToFrtIdir+"/Table6.xlsx")
	
	del _RED_MARROW[:]
	del _THYROID[:]
	del _LUNGS[:]
	del _SMALL_INTESTINE_WALL[:]
	del _SKIN[:]
	del _TESTES[:]
	del _OVARIES[:]
	del _EFFECTIVE_DOSE[:]
	
	return

# Таблица 7 – Прогнозируемые дозы за счет внешнего облучения от облака и от выпадений на поверхность земли и внутреннего облучения от ингаляционного поступления радионуклидов (взрослые), Зв . scenario
def createTable7(scenario):
	_THYROID_7 = sumGridForDoseNew("f200","f201","f173")
	_FOETUS_7 = sumGridForDoseNew("f3","f5","f2")
	_EFFECTIVE_DOSE_7 = sumGridForDoseNew("f198","f199","f141")
	_FOETUS_365 = sumGridForDoseNew("f3","f4","f2")
	_EFFECTIVE_DOSE_365 = sumGridForDoseNew("f204","f205","f141")
	
	wb = openpyxl.Workbook()
	ws = wb.worksheets[0]
	ws.title = u'Таблица 7'
	ws.merge_cells('A1:F1')
	
	ws['A1'] = u'Таблица 7 – Прогнозируемые дозы за счет внешнего облучения от облака и от выпадений на поверхность земли и внутреннего облучения от ингаляционного поступления радионуклидов (взрослые), Зв . scenario'.replace("scenario",scenario)
	ws['A2'] = u'Населенные пункты (долгота; широта)'
	
	ws['B2'] = u'Эквивалентная доза на щитовидную железу за 7 суток'
	ws['C2'] = u'Эквивалентная доза на плод за 7 суток'
	ws['D2'] = u'Эффективная доза за 7 суток'
	ws['E2'] = u'Эквивалентная доза на плод за первый год'
	ws['F2'] = u'Эффективная доза за первый год'
	
	curPos = 3
	for pointId, point in enumerate(pointsForAnalysis):
		ws['A'+str(curPos)] = point[0]+" ("+ format(point[1], '.6f').replace('.', ',')+";"+format(point[2], '.6f').replace('.', ',')+")"
		ws['B'+str(curPos)] = format(getValueForConfidence(pointId,_THYROID_7), '.2E').replace('.', ',')
		ws['C'+str(curPos)] = format(getValueForConfidence(pointId,_FOETUS_7), '.2E').replace('.', ',')
		ws['D'+str(curPos)] = format(getValueForConfidence(pointId,_EFFECTIVE_DOSE_7), '.2E').replace('.', ',')
		ws['E'+str(curPos)] = format(getValueForConfidence(pointId,_FOETUS_365) , '.2E').replace('.', ',')
		ws['F'+str(curPos)] = format(getValueForConfidence(pointId,_EFFECTIVE_DOSE_365), '.2E').replace('.', ',')
		curPos = curPos + 1
	
	wb.save(pathToFrtIdir+"/Table7.xlsx")
	
	del _THYROID_7[:]
	del _FOETUS_7[:]
	del _EFFECTIVE_DOSE_7[:]
	del _FOETUS_365[:]
	del _EFFECTIVE_DOSE_365[:]
	
	return

# Таблица 8 – Прогнозируемые дозы за счет внешнего облучения от облака и от выпадений на поверхность земли и внутреннего облучения от ингаляционного поступления радионуклидов (дети), Зв . scenario
def createTable8(scenario):
	_THYROID_7 =    sumGridForDose(sumGridForDoseNew("f200","f201"),maxGridForDoseNew("f168","f169","f170","f171","f172"))
	_EFFECTIVE_DOSE_7 =  sumGridForDose(sumGridForDoseNew("f198","f199"),maxGridForDoseNew("f136","f137","f138","f139","f140"))
	_EFFECTIVE_DOSE_365 = sumGridForDose(sumGridForDoseNew("f204","f205"),maxGridForDoseNew("f136","f137","f138","f139","f140"))
	
	wb = openpyxl.Workbook()
	ws = wb.worksheets[0]
	ws.title = u'Таблица 8'
	ws.merge_cells('A1:D1')
	
	ws['A1'] = u'Таблица 8 – Прогнозируемые дозы за счет внешнего облучения от облака и от выпадений на поверхность земли и внутреннего облучения от ингаляционного поступления радионуклидов (дети), Зв . scenario'.replace("scenario",scenario)
	ws['A2'] = u'Населенные пункты (долгота; широта)'
	
	ws['B2'] = u'Эквивалентная доза на щитовидную железу за 7 суток'
	ws['C2'] = u'Эффективная доза за 7 суток'
	ws['D2'] = u'Эффективная доза за первый год'
	
	curPos = 3
	for pointId, point in enumerate(pointsForAnalysis):
		ws['A'+str(curPos)] = point[0]+" ("+ format(point[1], '.6f').replace('.', ',')+";"+format(point[2], '.6f').replace('.', ',')+")"
		ws['B'+str(curPos)] = format(getValueForConfidence(pointId,_THYROID_7), '.2E').replace('.', ',')
		ws['C'+str(curPos)] = format(getValueForConfidence(pointId,_EFFECTIVE_DOSE_7), '.2E').replace('.', ',')
		ws['D'+str(curPos)] = format(getValueForConfidence(pointId,_EFFECTIVE_DOSE_365), '.2E').replace('.', ',')
		curPos = curPos + 1
	
	wb.save(pathToFrtIdir+"/Table8.xlsx")
	
	del _THYROID_7[:]
	del _EFFECTIVE_DOSE_7[:]
	del _EFFECTIVE_DOSE_365[:]
	
	return

# Таблица 9 – Прогнозируемые дозы за счет попадания радионуклидов внутрь организма пероральным путем, Зв . scenario 
def createTable9(scenario,milk_plants_meat = [0.0]*3):
	_FALLOUT = sumGridForDoseNew("f0")
	
	wb = openpyxl.Workbook()
	ws = wb.worksheets[0]
	ws.title = u'Таблица 9'
	ws.merge_cells('A1:D1')
	
	ws['A1'] = u'Таблица 9 – Прогнозируемые дозы за счет попадания радионуклидов внутрь организма пероральным путем, Зв . scenario'.replace("scenario",scenario)
	ws['A2'] = u'Населенные пункты (долгота; широта)'
	
	ws['B2'] = u'Эффективная доза за первый год за счет потребления молока'
	ws['C2'] = u'Эффективная доза за первый год за счет потребления мяса'
	ws['D2'] = u'Эффективная доза за первый год за счет потребления овощей'
	
	curPos = 3
	for pointId, point in enumerate(pointsForAnalysis):
		ws['A'+str(curPos)] = point[0]+" ("+ format(point[1], '.6f').replace('.', ',')+";"+format(point[2], '.6f').replace('.', ',')+")"
		ws['B'+str(curPos)] = format(getValueForConfidence(pointId,_FALLOUT)*milk_plants_meat[0], '.2E').replace('.', ',')
		ws['C'+str(curPos)] = format(getValueForConfidence(pointId,_FALLOUT)*milk_plants_meat[2], '.2E').replace('.', ',')
		ws['D'+str(curPos)] = format(getValueForConfidence(pointId,_FALLOUT)*milk_plants_meat[1], '.2E').replace('.', ',')
		curPos = curPos + 1
	
	wb.save(pathToFrtIdir+"/Table9.xlsx")
	
	del _FALLOUT[:]
	return

#Таблица 10 – Прогнозируемые плотность поверхностных выпадений и проинтегрированная по времени концентрация
def createTable10(scenario):
	_FALLOUT = sumGridForDoseNew("f0")
	_TIC = sumGridForDoseNew("f1")
	
	wb = openpyxl.Workbook()
	ws = wb.worksheets[0]
	ws.title = u'Таблица 10'
	ws.merge_cells('A1:C1')
	
	ws['A1'] = u'Таблица 10 – Прогнозируемые плотность поверхностных выпадений и проинтегрированная по времени концентрация. scenario'.replace("scenario",scenario)
	ws['A2'] = u'Населенные пункты (долгота; широта)'
	
	ws['B2'] = u'Плотность поверхностных выпадений, Бк/кв.м.'
	ws['C2'] = u'Проинтегрированная по времени концентрация, Бк·с/кб.м.'
	
	curPos = 3
	for pointId, point in enumerate(pointsForAnalysis):
		ws['A'+str(curPos)] = point[0]+" ("+ format(point[1], '.6f').replace('.', ',')+";"+format(point[2], '.6f').replace('.', ',')+")"
		ws['B'+str(curPos)] = format(getValueForConfidence(pointId,_FALLOUT), '.2E').replace('.', ',')
		ws['C'+str(curPos)] = format(getValueForConfidence(pointId,_TIC), '.2E').replace('.', ',')
		curPos = curPos + 1
	
	wb.save(pathToFrtIdir+"/Table10.xlsx")
	
	del _FALLOUT[:]
	del _TIC[:]
	return

def main():
	mypath=os.path.dirname(os.path.realpath( __file__ ))
	os.chdir(mypath)
	global pointsForAnalysis, arrayOfResultsForPoints
	pointsForAnalysis = reaNPFromFile(pathToNPfile)
	arrayOfResultsForPoints = [[] for x in pointsForAnalysis]
	print pointsForAnalysis
	myWorkPath = pathToFrtIdir
	make_sure_path_exists(myWorkPath)
	prepareArrayForVarSeries()
	
	#out = open(pathToFrtIdir+"/functionals.dat","wt")
	#pointId = 0
	#for i in range(len(arrayOfResultsForPoints[pointId])):
	#	out.write('f'+str(i)+'\t')
	#out.write('\r\n')
	#for j in range(len(arrayOfResultsForPoints[pointId][0])):
	#	for i in range(len(arrayOfResultsForPoints[pointId])):
	#		out.write('%0.12e' % arrayOfResultsForPoints[pointId][i][j]+'\t')
	#	out.write('\r\n')
	#out.close()
	scenario = u"xxx"
	createTable1(scenario)
	createTable2(scenario)
	createTable3(scenario)
	createTable4(scenario)
	createTable5(scenario)
	createTable6(scenario)
	createTable7(scenario)
	createTable8(scenario)
	createTable10(scenario)
	createTable9(scenario,milk_plants_meat = [1.79E-08, 6.31E-09, 9.67E-09])
	
	print("Ok!")
	return

if __name__ == "__main__":
	sys.exit(main())